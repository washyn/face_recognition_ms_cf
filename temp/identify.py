import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter, Cell, column_index_from_string
import time
from global_variables import URL
from global_variables import cfKey
###################################################################################
# prepare
#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')
###################################################################################
def getDateColumn():
	for i in range(1, len(sheet.rows[0]) + 1):
		col = get_column_letter(i)
		if sheet.cell('%s%s'% (col,'1')).value == currentDate:
			return col
			
###################################################################################

CF.Key.set(cfKey)
CF.BaseUrl.set(URL)



connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()
###################################################################################
# Asistencias
attend = [0 for i in range(60)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, '../Cropped_faces')
###################################################################################
# se probara con cada uno de los archivos, con varias fotos se trata de detectar la 
# asistencia
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.pathname2url(os.path.join(directory, filename))
		# se trata de detectar caras en la imangen
		res = CF.face.detect(imgurl)
		if len(res) != 1:
			print("No face detected.")
			continue
			
		faceIds = []
		# para una cara en el resultado de caras
		# se agrega en el array de Id de caras
		for face in res:
			faceIds.append(face['faceId'])
		
		# identificar en el grupo de personas
		res = CF.face.identify(faceIds, personGroupId)

		print(filename)
		print(res)
		
		# si es que esta en los candidatos se otiene el personId,
		# se busca en db  y se muestra toda su info
		for face  in res:
			if not face['candidates']:
				print("Unknown") 
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
				attend[int(row[0])] += 1
				print(row[1] + " recognized")
		time.sleep(6)
		
#############################################################################
# se establece los valores de asitencia para guardarlos
for row in range(2, len(sheet.columns[0]) + 1):
	rn = sheet.cell('A%s'% row).value
	if rn is not None:
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			sheet['%s%s' % (col, str(row))] = 1

# Se guarda el excel
wb.save(filename = "reports.xlsx")	 	

########################################################################
# en esta seccion se prueba con una soola imangen y se muestra sus datos

#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
#   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print(personName)
#print(res)
